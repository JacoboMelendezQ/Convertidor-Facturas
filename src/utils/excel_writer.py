"""
Escritor de Excel con formato colombiano.
- Precios: punto miles, coma decimal, máximo 2 decimales si los tiene
- Cantidades: entero sin decimales
- Porcentajes: número entero con %
- Encabezados con fondo azul claro, datos en blanco
- Fila TOTAL al final con fórmula SUM
- Fuente Arial 10pt, ancho automático
"""
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.utils.cleaner import classify_column, find_value_column, parse_number, parse_percentage

# Estilos
_HEADER_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
_TOTAL_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
_FONT = Font(name='Arial', size=10)
_FONT_BOLD = Font(name='Arial', size=10, bold=True)
_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
_ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
_ALIGN_LEFT = Alignment(horizontal='left', vertical='center')

# Formatos numéricos (funcionan con localización española — punto=miles, coma=decimal)
_FMT_PRICE = '#,##0.##'      # Muestra hasta 2 decimales, omite ceros trailing
_FMT_QTY = '#,##0'
_FMT_PCT = '0%'


def _to_number(val: str, col_type: str):
    """Convierte string a número según el tipo de columna."""
    if not val or not val.strip():
        return None
    if col_type == 'pct':
        pct = parse_percentage(val)
        return pct / 100.0 if pct != 0 else None
    num = parse_number(val)
    return num if num != 0 or re.search(r'\d', str(val)) else None


def write_excel(headers: list, rows: list[dict], output_path: str) -> str:
    """
    Genera un archivo Excel con los datos extraídos.
    Retorna la ruta del archivo creado.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Factura'

    if not headers:
        wb.save(output_path)
        return output_path

    # ── Detectar tipos de columna ──
    col_types = {h: classify_column(h) for h in headers}
    value_col = find_value_column(headers)

    # ── Fila de encabezados (fila 1) ──
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _FONT_BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTER

    # ── Filas de datos ──
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            raw = row.get(header, '')
            col_type = col_types[header]
            cell = ws.cell(row=row_idx, column=col_idx)

            # Porcentajes: guardar como texto "20%" para evitar problemas de locale.
            # Solo aplica si el valor contiene '%' o es puramente numérico.
            # Excepción: si el valor no lleva '%' y es >100, es un monto en pesos
            # (e.g. columna IVA con valor 199 pesos), no un porcentaje.
            if col_type == 'pct':
                raw_s = str(raw).strip()
                is_pct_value = '%' in raw_s or re.match(r'^[\d.,\s]+$', raw_s)
                if is_pct_value and re.search(r'\d', raw_s):
                    pct_val = parse_percentage(raw_s)
                    if '%' not in raw_s and pct_val > 100:
                        cell.value = int(pct_val) if pct_val == int(pct_val) else pct_val
                        cell.number_format = '#,##0'
                        cell.alignment = _ALIGN_RIGHT
                    else:
                        cell.value = f'{int(round(pct_val))}%'
                        cell.alignment = _ALIGN_RIGHT
                else:
                    cell.value = raw_s
                    cell.alignment = _ALIGN_LEFT
                cell.font = _FONT
                continue

            num = _to_number(raw, col_type)

            if num is not None and col_type != 'text':
                if col_type == 'price':
                    # Store as int when whole number (avoids 28315.0 display)
                    cell.value = int(num) if num == int(num) else num
                    cell.number_format = '#,##0' if num == int(num) else _FMT_PRICE
                    cell.alignment = _ALIGN_RIGHT
                elif col_type == 'qty':
                    cell.value = int(num) if num == int(num) else num
                    cell.number_format = _FMT_QTY
                    cell.alignment = _ALIGN_RIGHT
            else:
                val_str = raw.strip() if raw else ''
                cell.value = val_str
                cell.alignment = _ALIGN_LEFT
                # Long numeric codes → force text format to prevent scientific notation
                if re.match(r'^\d{10,}$', val_str):
                    cell.number_format = '@'

            cell.font = _FONT

    data_last_row = len(rows) + 1  # última fila de datos

    # ── Ancho automático de columnas ──
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row in rows:
            val = str(row.get(header, '') or '')
            if len(val) > max_len:
                max_len = len(val)
        max_len = min(max_len + 2, 60)
        ws.column_dimensions[col_letter].width = max(max_len, 10)

    # ── Altura de filas ──
    ws.row_dimensions[1].height = 20
    for i in range(2, data_last_row + 1):
        ws.row_dimensions[i].height = 16

    # Inmovilizar encabezado
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    return output_path
